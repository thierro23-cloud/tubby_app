# =============================================================================
# BLUEPRINT: Plan de Protección
# =============================================================================
#
# Archivo: blueprints/inventario/btn_plan_de_proteccion_form_bp.py
#
# Descripción:
#   Módulo de gestión completa (CRUD) de la tabla `tbl_plan_de_proteccion`
#   dentro de la base de datos `inventario`.
#
# Funcionalidad principal:
#   1. Listar planes de protección existentes.
#   2. Crear un nuevo plan de protección.
#   3. Editar un plan existente.
#   4. Ver detalle de un plan en modo solo lectura.
#   5. Eliminar un plan de protección.
#   6. Exportar listado a Excel.
#
# Integración:
#   - Base de datos: módulo `db.py` (ejecutar_query / ejecutar_non_query).
#   - Formularios: WTForms (Form, sin FlaskForm).
#   - Seguridad: decorador `login_required` de `services.helpers`.
#   - Plantilla: `plan_de_proteccion.html` con modos: lista / formulario / ver.
#
# Convención del panel super admin:
#   - Blueprint: btn_plan_de_proteccion_form_bp
#   - Módulo padre: modulo_inventario_plan_de_proteccion_bp
#   - Vista principal: btn_plan_de_proteccion_listado (empieza por "btn_")
#
# =============================================================================

# =============================================================================
# 1. IMPORTS
# =============================================================================

# 1.1 Flask core
from flask import (
    Blueprint,
    render_template,
    redirect,
    url_for,
    flash,
    request,
    send_file,
)

# 1.2 Seguridad
from services.helpers import login_required

# 1.3 Acceso a base de datos
from db import ejecutar_query, ejecutar_non_query

# 1.4 WTForms (validación de formularios sin FlaskForm)
from wtforms import (
    Form,
    StringField,
    IntegerField,
    DecimalField,
    TextAreaField,
    SelectField,
    DateField,
)
from wtforms.validators import DataRequired, Optional, Length

# 1.5 Sistema de archivos (para verificar existencia de imágenes)
import os

# 1.6 Excel: generación en memoria con openpyxl
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
# [web:128][web:134]


# =============================================================================
# 2. DEFINICIÓN DEL FORMULARIO
# =============================================================================

class PlanProteccionForm(Form):
    """
    Formulario para gestionar un registro de tbl_plan_de_proteccion.

    Notas:
      - Usa wtforms.Form en lugar de FlaskForm (sin dependencias externas).
      - Los nombres de campo coinciden con los alias del SELECT (minúsculas).
      - La vista se encarga de pasar request.form al constructor.
    """

    # 2.1 Campo ID (solo lectura, autoincremental)
    Idtbl_general_plan_prevencion = IntegerField(
        "ID",
        validators=[Optional()],
    )

    # 2.2 Relaciones con otras tablas (SelectField)
    Idtbl_edificios_municipales = SelectField(
        "inmueble",
        coerce=int,
        validators=[DataRequired(message="Selecciona un edificio")],
    )

    Idtbl_ref_edificio = SelectField(
        "Ref. edificio",
        coerce=int,
        validators=[Optional()],
    )

    idtbl_piezas_coleccion = SelectField(
        "Pieza de colección",
        coerce=int,
        validators=[Optional()],
    )

    idtbl_materia_tecnica = SelectField(
        "Materia técnica",
        coerce=int,
        validators=[Optional()],
    )

    idtbl_prioridad = SelectField(
        "Prioridad",
        coerce=int,
        validators=[Optional()],
    )

    idtbl_material_nesario = SelectField(
        "Material necesario",
        coerce=int,
        validators=[Optional()],
    )

    idtbl_ruta_evacuacion = SelectField(
        "Ruta de evacuación",
        coerce=int,
        validators=[Optional()],
    )

    idtbl_lugar_de_deposito = SelectField(
        "Lugar de depósito",
        coerce=int,
        validators=[Optional()],
    )

    idtbl_material_necesario_para_su_proteccion = SelectField(
        "Material para su protección",
        coerce=int,
        validators=[Optional()],
    )

    # 2.3 Campos de texto cortos
    numero_inventario = StringField(
        "Número de inventario",
        validators=[Optional(), Length(max=255)],
    )

    ubicacion = StringField(
        "Ubicación",
        validators=[Optional(), Length(max=255)],
    )

    dimensiones = StringField(
        "Dimensiones",
        validators=[Optional(), Length(max=255)],
    )

    dimensiones2 = StringField(
        "Dimensiones 2",
        validators=[Optional(), Length(max=255)],
    )

    estado = StringField(
        "Estado",
        validators=[Optional(), Length(max=255)],
    )

    estado_de_conservacion = StringField(
        "Estado de conservación",
        validators=[Optional(), Length(max=255)],
    )

    responsable_de_evacuacion = StringField(
        "Responsable de evacuación",
        validators=[Optional(), Length(max=255)],
    )

    vehiculo_empleado = StringField(
        "Vehículo empleado",
        validators=[Optional(), Length(max=255)],
    )

    responsable_del_deposito = StringField(
        "Responsable del depósito",
        validators=[Optional(), Length(max=255)],
    )

    conductor = StringField(
        "Conductor",
        validators=[Optional(), Length(max=255)],
    )

    # 2.4 Campos de texto largos
    descripcion = TextAreaField(
        "Descripción",
        validators=[Optional()],
    )

    observaciones = TextAreaField(
        "Observaciones",
        validators=[Optional()],
    )

    # 2.5 Campos numéricos
    numero_de_piezas = IntegerField(
        "Número de piezas",
        validators=[Optional()],
    )

    numero_personas_necesarias = IntegerField(
        "Personas necesarias",
        validators=[Optional()],
    )

    peso_aproximado = DecimalField(
        "Peso aproximado",
        validators=[Optional()],
        places=2,
    )

    # 2.6 Fechas
    fecha_de_inspeccion = DateField(
        "Fecha de inspección",
        format="%Y-%m-%d",
        validators=[Optional()],
    )

    proxima_fecha_inspeccion_recomendada = DateField(
        "Próxima fecha inspección recomendada",
        format="%Y-%m-%d",
        validators=[Optional()],
    )

    # 2.7 Archivos (nombres de fichero, no file upload aquí)
    foto = StringField(
        "Foto (nombre fichero)",
        validators=[Optional(), Length(max=255)],
    )

    plano = StringField(
        "Plano (nombre fichero)",
        validators=[Optional(), Length(max=255)],
    )


# =============================================================================
# 3. DEFINICIÓN DE LA BLUEPRINT
# =============================================================================

btn_plan_de_proteccion_form_bp = Blueprint(
    "btn_plan_de_proteccion_form_bp",
    __name__,
    url_prefix="/plan_de_proteccion",
)


# =============================================================================
# 4. RUTA: LISTADO DE PLANES DE PROTECCIÓN
# =============================================================================

@btn_plan_de_proteccion_form_bp.route("/")
@login_required
def btn_plan_de_proteccion_listado():
    sql = """
        SELECT
            Idtbl_general_plan_prevencion,
            idtbl_edificio,
            Idtbl_ref_edificio,
            idtbl_piezas_coleccion,
            numero_de_inventario AS numero_inventario,
            `Descripción` AS descripcion,
            idtbl_materia_tecnica,
            ubicacion,
            idtbl_prioridad,
            Dimensiones AS dimensiones,
            Dimensiones2 AS dimensiones2,
            numero_de_piezas,
            Estado AS estado,
            numero_personas_necesarias,
            idtbl_material_nesario,
            idtbl_ruta_evacuacion,
            idtbl_lugar_de_deposito,
            idtbl_material_necesario_para_su_proteccion,
            fecha_de_inspeccion,
            estado_de_conservacion,
            proxima_fecha_inspeccion_recomendada,
            peso_aproximado,
            responsable_de_evacuacion,
            vehiculo_empleado,
            responsable_del_deposito,
            Conductor AS conductor,
            Observaciones AS observaciones
        FROM tbl_plan_de_proteccion
        ORDER BY Idtbl_general_plan_prevencion DESC
    """

    planes = ejecutar_query(sql, nombre_bd="inventario")

    # Postprocesar descripciones para el listado: crear campo descripcion_corta
    for plan in planes:
        desc = plan.get("descripcion")
        if desc is None:
            plan["descripcion_corta"] = ""
        else:
            desc_str = str(desc)
            plan["descripcion_corta"] = (
                desc_str[:50] + "..." if len(desc_str) > 50 else desc_str
            )

    return render_template(
        "plan_de_proteccion.html",
        modo="lista",
        planes=planes,
    )


# =============================================================================
# 4.1 RUTA: EXPORTAR LISTADO A EXCEL
# =============================================================================

@btn_plan_de_proteccion_form_bp.route("/exportar_excel", methods=["GET"])
@login_required
def exportar_excel():
    """
    4.1.1 Función: Exportar a Excel el listado de planes de protección.

    - Reutiliza el mismo origen de datos que el listado.
    - Genera un archivo XLSX en memoria con openpyxl.
    - Devuelve el fichero como attachment.
    """

    sql = """
        SELECT
            Idtbl_general_plan_prevencion,
            numero_de_inventario AS numero_inventario,
            `Descripción` AS descripcion,
            ubicacion,
            idtbl_prioridad,
            Estado AS estado
        FROM tbl_plan_de_proteccion
        ORDER BY Idtbl_general_plan_prevencion DESC
    """

    planes = ejecutar_query(sql, nombre_bd="inventario")

    # Crear libro y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Planes de protección"

    # Cabeceras y campos a exportar
    columnas = [
        ("ID", "Idtbl_general_plan_prevencion"),
        ("Nº Inventario", "numero_inventario"),
        ("Descripción", "descripcion"),
        ("Ubicación", "ubicacion"),
        ("Prioridad", "idtbl_prioridad"),
        ("Estado", "estado"),
    ]

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4B5563")  # gris oscuro
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # Escribir cabeceras
    for col_index, (titulo, _) in enumerate(columnas, start=1):
        cell = ws.cell(row=1, column=col_index, value=titulo)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Escribir datos
    row_index = 2
    for plan in planes:
        for col_index, (_, key) in enumerate(columnas, start=1):
            valor = plan.get(key)
            ws.cell(row=row_index, column=col_index, value=valor)
        row_index += 1

    # Ajustar anchos de columna
    ancho_sugerido = {
        "A": 10,   # ID
        "B": 15,   # Nº Inventario
        "C": 60,   # Descripción
        "D": 30,   # Ubicación
        "E": 12,   # Prioridad
        "F": 18,   # Estado
    }
    for col_letter, width in ancho_sugerido.items():
        ws.column_dimensions[col_letter].width = width

    # Congelar cabecera
    ws.freeze_panes = "A2"

    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "planes_proteccion.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
# [web:125][web:130]

# =============================================================================
# 4.2 RUTA: FORMULARIO FILTRADO POR EDIFICIO/OBRA (ENTRADA SUPER ADMIN)
# =============================================================================

@btn_plan_de_proteccion_form_bp.route("/formulario_filtrado", methods=["GET", "POST"])
@login_required
def btn_plan_de_proteccion_form_editar_bp():
    """
    4.2 Función: Entrada al formulario de planes con filtro por edificio/obra.

    - GET: Muestra un select de edificio y, opcionalmente, obra (cuando lo tengas),
      más una lista de planes de ese edificio para elegir y editar.
    - POST: según lo que envíes, redirige a 'editar' del id elegido
      o a 'nuevo' con el edificio preseleccionado.
    """

    # Cargar edificios
    filas_edificios = ejecutar_query(
        """
        SELECT Idtbl_edificios_municipales, inmueble
        FROM tbl_edificios_municipales
        ORDER BY inmueble
        """,
        nombre_bd="inventario",
    )

    edificios = [
        {"id": fila["Idtbl_edificios_municipales"], "nombre": fila["inmueble"]}
        for fila in filas_edificios
    ]

    id_edificio = request.args.get("id_edificio", type=int)
    planes = []

    if id_edificio:
        # Filtrar planes por edificio
        sql = """
            SELECT
                Idtbl_general_plan_prevencion,
                numero_de_inventario AS numero_inventario,
                `Descripción` AS descripcion,
                ubicacion,
                idtbl_prioridad,
                Estado AS estado
            FROM tbl_plan_de_proteccion
            WHERE idtbl_edificio = %s
            ORDER BY Idtbl_general_plan_prevencion DESC
        """
        planes = ejecutar_query(sql, (id_edificio,), nombre_bd="inventario")

    # Si se envía por POST podemos decidir qué hacer:
    if request.method == "POST":
        accion = request.form.get("accion")
        id_plan = request.form.get("id_plan", type=int)
        id_edificio_sel = request.form.get("id_edificio", type=int)

        if accion == "editar" and id_plan:
            return redirect(
                url_for("btn_plan_de_proteccion_form_bp.editar", id=id_plan)
            )
        if accion == "nuevo" and id_edificio_sel:
            # Pasar edificio seleccionado como parámetro para pre-cargar el form
            return redirect(
                url_for("btn_plan_de_proteccion_form_bp.nuevo", id_edificio=id_edificio_sel)
            )

    return render_template(
        "plan_de_proteccion_filtrado.html",
        edificios=edificios,
        id_edificio=id_edificio,
        planes=planes,
    )
# =============================================================================
# 5. RUTA: NUEVO PLAN DE PROTECCIÓN
# =============================================================================

@btn_plan_de_proteccion_form_bp.route("/nuevo", methods=["GET", "POST"])
@login_required
def nuevo():
    """
    5.1 Función: Crear un nuevo plan de protección.
    """

    form = PlanProteccionForm(request.form)
    cargar_opciones_plan_proteccion(form)

    if request.method == "POST" and form.validate():
        datos = form.data

        sql = """
            INSERT INTO tbl_plan_de_proteccion (
                idtbl_edificio,
                Idtbl_ref_edificio,
                idtbl_piezas_coleccion,
                numero_de_inventario,
                `Descripción`,
                idtbl_materia_tecnica,
                ubicacion,
                idtbl_prioridad,
                Dimensiones,
                Dimensiones2,
                numero_de_piezas,
                Estado,
                numero_personas_necesarias,
                idtbl_material_nesario,
                idtbl_ruta_evacuacion,
                idtbl_lugar_de_deposito,
                idtbl_material_necesario_para_su_proteccion,
                fecha_de_inspeccion,
                estado_de_conservacion,
                proxima_fecha_inspeccion_recomendada,
                peso_aproximado,
                responsable_de_evacuacion,
                vehiculo_empleado,
                responsable_del_deposito,
                Conductor,
                Observaciones
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s, %s
            )
        """

        params = (
            datos.get("idtbl_edificio"),
            datos.get("Idtbl_ref_edificio"),
            datos.get("idtbl_piezas_coleccion"),
            datos.get("numero_inventario"),
            datos.get("descripcion"),
            datos.get("idtbl_materia_tecnica"),
            datos.get("ubicacion"),
            datos.get("idtbl_prioridad"),
            datos.get("dimensiones"),
            datos.get("dimensiones2"),
            datos.get("numero_de_piezas"),
            datos.get("estado"),
            datos.get("numero_personas_necesarias"),
            datos.get("idtbl_material_nesario"),
            datos.get("idtbl_ruta_evacuacion"),
            datos.get("idtbl_lugar_de_deposito"),
            datos.get("idtbl_material_necesario_para_su_proteccion"),
            datos.get("fecha_de_inspeccion"),
            datos.get("estado_de_conservacion"),
            datos.get("proxima_fecha_inspeccion_recomendada"),
            datos.get("peso_aproximado"),
            datos.get("responsable_de_evacuacion"),
            datos.get("vehiculo_empleado"),
            datos.get("responsable_del_deposito"),
            datos.get("conductor"),
            datos.get("observaciones"),
        )

        ejecutar_non_query(sql, params, nombre_bd="inventario")

        fila_id = ejecutar_query(
            "SELECT MAX(Idtbl_general_plan_prevencion) AS id FROM tbl_plan_de_proteccion",
            nombre_bd="inventario",
        )
        nuevo_id = fila_id[0]["id"] if fila_id else None

        flash(
            f"Plan de protección creado correctamente "
            f"{'(ID: ' + str(nuevo_id) + ')' if nuevo_id is not None else ''}",
            "success",
        )

        if nuevo_id is not None:
            return redirect(
                url_for("btn_plan_de_proteccion_form_bp.ver", id=nuevo_id)
            )
        return redirect(
            url_for("btn_plan_de_proteccion_form_bp.btn_plan_de_proteccion_listado")
        )

    return render_template(
        "plan_de_proteccion.html",
        modo="formulario",
        form=form,
        plan=None,
        foto_url=None,
        plano_url=None,
    )


# =============================================================================
# 6. RUTA: EDITAR PLAN DE PROTECCIÓN
# =============================================================================

@btn_plan_de_proteccion_form_bp.route("/editar/<int:id>", methods=["GET", "POST"])
@login_required
def editar(id):
    """
    6.1 Función: Editar un plan de protección existente.
    """

    plan = _obtener_plan_por_id(id)
    if not plan:
        flash("Plan de protección no encontrado", "danger")
        return redirect(
            url_for("btn_plan_de_proteccion_form_bp.btn_plan_de_proteccion_listado")
        )

    if request.method == "POST":
        form = PlanProteccionForm(request.form)
    else:
        form = PlanProteccionForm(data=plan)

    cargar_opciones_plan_proteccion(form)

    if request.method == "POST" and form.validate():
        datos = form.data

        sql = """
            UPDATE tbl_plan_de_proteccion
            SET
                idtbl_edificio = %s,
                Idtbl_ref_edificio = %s,
                idtbl_piezas_coleccion = %s,
                numero_de_inventario = %s,
                `Descripción` = %s,
                idtbl_materia_tecnica = %s,
                ubicacion = %s,
                idtbl_prioridad = %s,
                Dimensiones = %s,
                Dimensiones2 = %s,
                numero_de_piezas = %s,
                Estado = %s,
                numero_personas_necesarias = %s,
                idtbl_material_nesario = %s,
                idtbl_ruta_evacuacion = %s,
                idtbl_lugar_de_deposito = %s,
                idtbl_material_necesario_para_su_proteccion = %s,
                fecha_de_inspeccion = %s,
                estado_de_conservacion = %s,
                proxima_fecha_inspeccion_recomendada = %s,
                peso_aproximado = %s,
                responsable_de_evacuacion = %s,
                vehiculo_empleado = %s,
                responsable_del_deposito = %s,
                Conductor = %s,
                Observaciones = %s
            WHERE Idtbl_general_plan_prevencion = %s
        """

        params = (
            datos.get("idtbl_edificio"),
            datos.get("Idtbl_ref_edificio"),
            datos.get("idtbl_piezas_coleccion"),
            datos.get("numero_inventario"),
            datos.get("descripcion"),
            datos.get("idtbl_materia_tecnica"),
            datos.get("ubicacion"),
            datos.get("idtbl_prioridad"),
            datos.get("dimensiones"),
            datos.get("dimensiones2"),
            datos.get("numero_de_piezas"),
            datos.get("estado"),
            datos.get("numero_personas_necesarias"),
            datos.get("idtbl_material_nesario"),
            datos.get("idtbl_ruta_evacuacion"),
            datos.get("idtbl_lugar_de_deposito"),
            datos.get("idtbl_material_necesario_para_su_proteccion"),
            datos.get("fecha_de_inspeccion"),
            datos.get("estado_de_conservacion"),
            datos.get("proxima_fecha_inspeccion_recomendada"),
            datos.get("peso_aproximado"),
            datos.get("responsable_de_evacuacion"),
            datos.get("vehiculo_empleado"),
            datos.get("responsable_del_deposito"),
            datos.get("conductor"),
            datos.get("observaciones"),
            id,
        )

        ejecutar_non_query(sql, params, nombre_bd="inventario")

        flash("Plan de protección actualizado correctamente", "success")

        return redirect(url_for("btn_plan_de_proteccion_form_bp.ver", id=id))

    foto_url, plano_url = _construir_rutas_imagenes(plan)

    return render_template(
        "plan_de_proteccion.html",
        modo="formulario",
        form=form,
        plan=plan,
        foto_url=foto_url,
        plano_url=plano_url,
    )


# =============================================================================
# 7. RUTA: VER PLAN DE PROTECCIÓN (SOLO LECTURA)
# =============================================================================

@btn_plan_de_proteccion_form_bp.route("/ver/<int:id>")
@login_required
def ver(id):
    """
    7.1 Función: Ver detalle de un plan de protección (solo lectura).
    """

    plan = _obtener_plan_por_id(id)
    if not plan:
        flash("Plan de protección no encontrado", "danger")
        return redirect(
            url_for("btn_plan_de_proteccion_form_bp.btn_plan_de_proteccion_listado")
        )

    foto_url, plano_url = _construir_rutas_imagenes(plan)

    return render_template(
        "plan_de_proteccion.html",
        modo="ver",
        plan=plan,
        form=None,
        foto_url=foto_url,
        plano_url=plano_url,
    )


# =============================================================================
# 8. RUTA: ELIMINAR PLAN DE PROTECCIÓN
# =============================================================================

@btn_plan_de_proteccion_form_bp.route("/eliminar/<int:id>", methods=["POST"])
@login_required
def eliminar(id):
    """
    8.1 Función: Eliminar un plan de protección.
    """

    sql = """
        DELETE FROM tbl_plan_de_proteccion
        WHERE Idtbl_general_plan_prevencion = %s
    """

    ejecutar_non_query(sql, (id,), nombre_bd="inventario")

    flash("Plan de protección eliminado correctamente", "success")

    return redirect(
        url_for("btn_plan_de_proteccion_form_bp.btn_plan_de_proteccion_listado")
    )


# =============================================================================
# 9. FUNCIÓN AUXILIAR: CARGAR OPCIONES DE SELECTFIELD
# =============================================================================

def cargar_opciones_plan_proteccion(form: PlanProteccionForm) -> None:
    """
    9.1 Función: Cargar opciones de los SelectField desde la base de datos.
    """

    filas_edificios = ejecutar_query(
        """
        SELECT Idtbl_edificios_municipales, inmueble
        FROM tbl_edificios_municipales
        ORDER BY inmueble
        """,
        nombre_bd="inventario",
    )

    # Opción por defecto + lista real
    form.idtbl_edificio.choices = [(0, "-- Seleccionar --")] + [
        (fila["Idtbl_edificios_municipales"], fila["inmueble"])
        for fila in filas_edificios
    ]

    # El resto sigue como placeholder
    form.Idtbl_ref_edificio.choices = [(0, "-- Pendiente configurar --")]
    form.idtbl_piezas_coleccion.choices = [(0, "-- Pendiente configurar --")]
    form.idtbl_materia_tecnica.choices = [(0, "-- Pendiente configurar --")]
    form.idtbl_prioridad.choices = [(0, "-- Pendiente configurar --")]
    form.idtbl_material_nesario.choices = [(0, "-- Pendiente configurar --")]
    form.idtbl_ruta_evacuacion.choices = [(0, "-- Pendiente configurar --")]
    form.idtbl_lugar_de_deposito.choices = [(0, "-- Pendiente configurar --")]
    form.idtbl_material_necesario_para_su_proteccion.choices = [
        (0, "-- Pendiente configurar --")
    ]


# =============================================================================
# 10. FUNCIÓN AUXILIAR: OBTENER PLAN POR ID
# =============================================================================

def _obtener_plan_por_id(id_plan: int) -> dict | None:
    """
    10.1 Función: Recuperar un único plan de protección por su ID.
    """

    sql = """
        SELECT
            Idtbl_general_plan_prevencion,
            idtbl_edificio,
            Idtbl_ref_edificio,
            idtbl_piezas_coleccion,
            numero_de_inventario AS numero_inventario,
            `Descripción` AS descripcion,
            idtbl_materia_tecnica,
            ubicacion,
            idtbl_prioridad,
            Dimensiones AS dimensiones,
            Dimensiones2 AS dimensiones2,
            numero_de_piezas,
            Estado AS estado,
            numero_personas_necesarias,
            idtbl_material_nesario,
            idtbl_ruta_evacuacion,
            idtbl_lugar_de_deposito,
            idtbl_material_necesario_para_su_proteccion,
            fecha_de_inspeccion,
            estado_de_conservacion,
            proxima_fecha_inspeccion_recomendada,
            peso_aproximado,
            responsable_de_evacuacion,
            vehiculo_empleado,
            responsable_del_deposito,
            Conductor AS conductor,
            Observaciones AS observaciones
        FROM tbl_plan_de_proteccion
        WHERE Idtbl_general_plan_prevencion = %s
    """

    filas = ejecutar_query(sql, (id_plan,), nombre_bd="inventario")
    return filas[0] if filas else None


# =============================================================================
# 11. FUNCIÓN AUXILIAR: CONSTRUIR RUTAS DE IMÁGENES (FOTO Y PLANO)
# =============================================================================

def _construir_rutas_imagenes(plan: dict) -> tuple[str | None, str | None]:
    """
    11.1 Función: Construir las URLs de las imágenes de foto y plano.
    """

    foto_url = None
    plano_url = None

    id_plan = plan["Idtbl_general_plan_prevencion"]

    # Foto
    if plan.get("foto"):
        foto_url = url_for("static", filename=f"imagen/plan_proteccion/{plan['foto']}")
    else:
        foto_path = os.path.join(
            "static", "imagen", "plan_proteccion", f"Foto_ID_{id_plan}_1.jpg"
        )
        if os.path.exists(foto_path):
            foto_url = url_for(
                "static",
                filename=f"imagen/plan_proteccion/Foto_ID_{id_plan}_1.jpg",
            )

    # Plano
    if plan.get("plano"):
        plano_url = url_for("static", filename=f"imagen/plan_proteccion/{plan['plano']}")
    else:
        plano_path = os.path.join(
            "static", "imagen", "plan_proteccion", f"Plano_ID_{id_plan}_1.jpg"
        )
        if os.path.exists(plano_path):
            plano_url = url_for(
                "static",
                filename=f"imagen/plan_proteccion/Plano_ID_{id_plan}_1.jpg",
            )

    return foto_url, plano_url